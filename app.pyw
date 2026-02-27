import sys
import os
import json
import uuid
import math
import unicodedata
import xml.etree.ElementTree as ET
from PyQt6.QtWidgets import (QApplication, QMainWindow, QGraphicsScene, QGraphicsView, 
                             QGraphicsPathItem, QGraphicsLineItem, QGraphicsTextItem, 
                             QToolBar, QFileDialog, QInputDialog, QMessageBox, QGraphicsItemGroup,
                             QGraphicsItem, QGraphicsEllipseItem, QColorDialog, QLabel, QWidget, QStyle,
                             QDialog, QVBoxLayout, QHBoxLayout, QGroupBox, QRadioButton, QComboBox, QDoubleSpinBox, QPushButton)
from PyQt6.QtCore import Qt, QRectF, QPointF, QLineF, QMarginsF
from PyQt6.QtGui import (QPen, QBrush, QColor, QPainter, QImage, QPainterPath, 
                         QTransform, QPainterPathStroker, QAction, QActionGroup,
                         QPageSize, QPageLayout, QUndoStack, QUndoCommand, QCursor)
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewWidget, QPrinterInfo
from PyQt6.QtSvg import QSvgGenerator

import qtawesome as qta
import qdarktheme
import networkx as nx
import openpyxl
import ezdxf

GRID_SIZE = 20

class SceneStateCommand(QUndoCommand):
    def __init__(self, main_window, old_state, new_state, description):
        super().__init__(description)
        self.main_window = main_window
        self.old_state = old_state
        self.new_state = new_state
        self.is_first_redo = True 

    def undo(self):
        self.main_window.load_scene_json(self.old_state, clear_scene=True, generate_new_ids=False, is_undo_redo=True)

    def redo(self):
        if self.is_first_redo:
            self.is_first_redo = False
            return
        self.main_window.load_scene_json(self.new_state, clear_scene=True, generate_new_ids=False, is_undo_redo=True)


class FlowchartView(QGraphicsView):
    def __init__(self, scene):
        super().__init__(scene)
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setTransformationAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.zoom_factor = 1.15
        self.setDragMode(QGraphicsView.DragMode.RubberBandDrag)
        self.setRubberBandSelectionMode(Qt.ItemSelectionMode.IntersectsItemShape)
        self.setMouseTracking(True) 

    def wheelEvent(self, event):
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            if event.angleDelta().y() > 0: self.scale(self.zoom_factor, self.zoom_factor)
            else: self.scale(1 / self.zoom_factor, 1 / self.zoom_factor)
        else:
            super().wheelEvent(event)

    def leaveEvent(self, event):
        if hasattr(self.scene(), 'hide_preview_node'):
            self.scene().hide_preview_node()
        super().leaveEvent(event)


class NodeItem(QGraphicsPathItem):
    def __init__(self, x, y, text="Node", node_type="process", node_id=None, bg_color="#E1F5FE", text_color="#000000"):
        super().__init__()
        self.node_type = node_type
        self.node_id = node_id if node_id else str(uuid.uuid4())
        self.edges = []
        self.bg_color = QColor(bg_color)
        self.text_color = QColor(text_color)
        self.default_pen = QPen(Qt.GlobalColor.black, 2)
        self._is_highlighted = False
        self.orig_x = x
        self.orig_y = y

        path = QPainterPath()
        if self.node_type == "process": path.addRect(QRectF(-50, -25, 100, 50))
        elif self.node_type == "decision": path.moveTo(0,-35); path.lineTo(60,0); path.lineTo(0,35); path.lineTo(-60,0); path.closeSubpath()
        elif self.node_type == "data": path.moveTo(-35,-25); path.lineTo(65,-25); path.lineTo(35,25); path.lineTo(-65,25); path.closeSubpath()
        elif self.node_type == "terminal": path.addRoundedRect(QRectF(-50, -25, 100, 50), 25, 25)
        else: path.addRect(QRectF(-50, -25, 100, 50))

        self.setPath(path)
        self.setPos(x, y)
        self.setBrush(QBrush(self.bg_color))
        self.setPen(self.default_pen)
        self.setFlags(QGraphicsItem.GraphicsItemFlag.ItemIsMovable | QGraphicsItem.GraphicsItemFlag.ItemIsSelectable | QGraphicsItem.GraphicsItemFlag.ItemSendsGeometryChanges)
        
        self.text_item = QGraphicsTextItem(text)
        self.text_item.setParentItem(self)
        self.text_item.setDefaultTextColor(self.text_color)
        self.set_text(text)

    def _update_text_pos(self):
        r = self.boundingRect(); tr = self.text_item.boundingRect()
        self.text_item.setPos(r.center().x() - tr.width()/2, r.center().y() - tr.height()/2)

    def set_text(self, text):
        self.text_item.setHtml(f"<div align='center'>{text.replace(chr(10), '<br>')}</div>")
        self._update_text_pos()

    def set_bg_color(self, color: QColor):
        self.bg_color = color; self.setBrush(QBrush(self.bg_color))

    def set_text_color(self, color: QColor):
        self.text_color = color; self.text_item.setDefaultTextColor(self.text_color)

    def set_highlight(self, active: bool):
        if self._is_highlighted != active: self._is_highlighted = active; self.update() 

    def add_edge(self, edge):
        self.edges.append(edge)

    def itemChange(self, change, value):
        if change == QGraphicsItem.GraphicsItemChange.ItemPositionChange and self.scene():
            return QPointF(round(value.x()/GRID_SIZE)*GRID_SIZE, round(value.y()/GRID_SIZE)*GRID_SIZE)
        elif change == QGraphicsItem.GraphicsItemChange.ItemPositionHasChanged:
            for edge in self.edges: edge.update_position()
        return super().itemChange(change, value)

    def paint(self, painter, option, widget=None):
        if self._is_highlighted: painter.setPen(QPen(QColor("#FF5722"), 3, Qt.PenStyle.DashLine))
        elif self.isSelected(): painter.setPen(QPen(QColor("#3B82F6"), 3))
        else: painter.setPen(self.default_pen)
        painter.setBrush(self.brush()); painter.drawPath(self.path())

    def mouseDoubleClickEvent(self, event):
        new_text, ok = QInputDialog.getMultiLineText(None, "テキスト編集", "ノード名:", self.text_item.toPlainText())
        if ok: self.set_text(new_text); self.scene().main_window.push_undo_state("テキスト変更")
        super().mouseDoubleClickEvent(event)


class WaypointItem(QGraphicsEllipseItem):
    def __init__(self, x, y, edge):
        super().__init__(-6, -6, 12, 12)
        self.edge = edge
        self.setPos(x, y)
        self.orig_x = x
        self.orig_y = y
        self.setBrush(QBrush(QColor("#FF9800"))); self.setPen(QPen(Qt.GlobalColor.white, 2))
        self.setFlags(QGraphicsItem.GraphicsItemFlag.ItemIsMovable | QGraphicsItem.GraphicsItemFlag.ItemSendsGeometryChanges | QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)
        self.setZValue(1)

    def itemChange(self, change, value):
        if change == QGraphicsItem.GraphicsItemChange.ItemPositionChange and self.scene():
            return QPointF(round(value.x()/GRID_SIZE)*GRID_SIZE, round(value.y()/GRID_SIZE)*GRID_SIZE)
        elif change == QGraphicsItem.GraphicsItemChange.ItemPositionHasChanged:
            self.edge.update_position()
        return super().itemChange(change, value)

    def paint(self, painter, option, widget=None):
        painter.setPen(QPen(QColor("#3B82F6"), 2) if self.isSelected() else self.pen())
        painter.setBrush(self.brush()); painter.drawEllipse(self.rect())

    def mouseDoubleClickEvent(self, event):
        self.edge.remove_waypoint(self); super().mouseDoubleClickEvent(event)
        
    def mouseReleaseEvent(self, event):
        super().mouseReleaseEvent(event); self.ungrabMouse(); self.edge.check_waypoint_straightness(self)


class EdgeTextItem(QGraphicsTextItem):
    def __init__(self, text, edge):
        super().__init__(text)
        self.edge = edge
        self.setFlags(QGraphicsItem.GraphicsItemFlag.ItemIsMovable | QGraphicsItem.GraphicsItemFlag.ItemIsSelectable | QGraphicsItem.GraphicsItemFlag.ItemSendsGeometryChanges)
        self.setParentItem(edge); self.setDefaultTextColor(QColor("#333333"))
        self.manual_offset = None; self._is_dragging = False

    def mousePressEvent(self, event): self._is_dragging = True; super().mousePressEvent(event)
    def mouseReleaseEvent(self, event): self._is_dragging = False; super().mouseReleaseEvent(event)

    def itemChange(self, change, value):
        if change == QGraphicsItem.GraphicsItemChange.ItemPositionChange and self._is_dragging:
            base_pos = self.edge.get_auto_text_pos()
            if base_pos is not None: self.manual_offset = value - base_pos
        return super().itemChange(change, value)

    def mouseDoubleClickEvent(self, event):
        new_text, ok = QInputDialog.getMultiLineText(None, "エッジのテキスト編集", "線上のテキスト:", self.edge.raw_text)
        if ok: 
            self.edge.set_text(new_text)
            if self.scene() and hasattr(self.scene(), 'main_window'):
                self.scene().main_window.push_undo_state("エッジテキスト変更")

    def paint(self, painter, option, widget=None):
        option.state &= ~QStyle.StateFlag.State_Selected
        super().paint(painter, option, widget)
        if self.isSelected():
            painter.setPen(QPen(QColor("#3B82F6"), 1, Qt.PenStyle.DashLine)); painter.setBrush(Qt.BrushStyle.NoBrush); painter.drawRect(self.boundingRect())


class EdgeItem(QGraphicsPathItem):
    def __init__(self, source_node, target_node, label="", width=2, style="solid", routing="straight"):
        super().__init__()
        self.source_node = source_node
        self.target_node = target_node
        self.raw_text = label
        self.waypoints = []
        self._drag_start_pos = None; self._potential_waypoint_index = -1
        
        self.line_width = width
        self.line_style = style
        self.routing = routing
        self.update_pen()
        
        self.setZValue(-1); self.setFlags(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)
        self.text_item = EdgeTextItem("", self)
        self._set_label_html(label)
        self.update_position()

    def update_pen(self):
        style_map = {"solid": Qt.PenStyle.SolidLine, "dash": Qt.PenStyle.DashLine, "dot": Qt.PenStyle.DotLine}
        ps = style_map.get(self.line_style, Qt.PenStyle.SolidLine)
        self.default_pen = QPen(Qt.GlobalColor.black, self.line_width, ps)
        self.setPen(self.default_pen)

    def boundingRect(self): return super().boundingRect().adjusted(-10, -10, 10, 10)
    def shape(self): stroker = QPainterPathStroker(); stroker.setWidth(20); return stroker.createStroke(super().shape())

    def _set_label_html(self, text):
        self.raw_text = text
        if text: 
            self.text_item.setHtml(f"<div style='font-weight: bold; font-family: sans-serif; text-align: center;'>{text.replace(chr(10), '<br>')}</div>")
            self.text_item.show()
        else: 
            self.text_item.setHtml("")
            self.text_item.hide()

    def set_text(self, text): 
        self._set_label_html(text)
        self.update_position()

    def get_auto_text_pos(self):
        if not self.source_node or not self.target_node: return None
        path = self.path()
        if path.isEmpty(): return QPointF(0, 0)
        
        # どのような線（直線・直角）でも、パスの長さのちょうど半分の位置にテキストを配置
        c = path.pointAtPercent(0.5)
        r = self.text_item.boundingRect()
        return QPointF(c.x() - r.width()/2, c.y() - r.height()/2 - 15)

    def _get_orthogonal_path(self, p1, p2):
        path = QPainterPath(); path.moveTo(p1)
        mid_y = (p1.y() + p2.y()) / 2
        path.lineTo(p1.x(), mid_y)
        path.lineTo(p2.x(), mid_y)
        path.lineTo(p2)
        return path

    def update_position(self):
        if not self.source_node or not self.target_node: return
        self.prepareGeometryChange()
        
        pts = [self.source_node.scenePos()] + [wp.scenePos() for wp in self.waypoints] + [self.target_node.scenePos()]
        path = QPainterPath()
        
        if self.routing == "orthogonal" and not self.waypoints:
            path = self._get_orthogonal_path(pts[0], pts[-1])
        else:
            path.moveTo(pts[0])
            for i in range(1, len(pts)):
                if self.routing == "orthogonal":
                    mid_y = (pts[i-1].y() + pts[i].y()) / 2
                    path.lineTo(pts[i-1].x(), mid_y)
                    path.lineTo(pts[i].x(), mid_y)
                path.lineTo(pts[i])

        self.setPath(path)
        if self.raw_text:
            base = self.get_auto_text_pos()
            if base is not None: self.text_item.setPos(base + self.text_item.manual_offset if self.text_item.manual_offset else base)

    def paint(self, painter, option, widget=None):
        painter.setPen(QPen(QColor("#3B82F6"), max(3, self.line_width)) if self.isSelected() else self.default_pen)
        painter.drawPath(self.path())
        if self.isSelected():
            painter.setPen(Qt.PenStyle.NoPen); painter.setBrush(QBrush(QColor("#3B82F6")))
            pts = [self.source_node.scenePos()] + [wp.scenePos() for wp in self.waypoints] + [self.target_node.scenePos()]
            for i in range(len(pts)-1):
                painter.drawEllipse(QPointF((pts[i].x()+pts[i+1].x())/2, (pts[i].y()+pts[i+1].y())/2), 5.0, 5.0)

    def mousePressEvent(self, event):
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier: super().mousePressEvent(event); return
        if event.button() == Qt.MouseButton.LeftButton and self.routing != "orthogonal":
            pos = event.scenePos()
            pts = [self.source_node.scenePos()] + [wp.scenePos() for wp in self.waypoints] + [self.target_node.scenePos()]
            for i in range(len(pts) - 1):
                if math.hypot(pos.x() - (pts[i].x()+pts[i+1].x())/2, pos.y() - (pts[i].y()+pts[i+1].y())/2) < 30:
                    self._drag_start_pos = pos; self._potential_waypoint_index = i; event.accept(); return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_start_pos and (event.scenePos() - self._drag_start_pos).manhattanLength() > 5:
            pos = event.scenePos()
            sx, sy = round(pos.x()/GRID_SIZE)*GRID_SIZE, round(pos.y()/GRID_SIZE)*GRID_SIZE
            wp = WaypointItem(sx, sy, self); self.waypoints.insert(self._potential_waypoint_index, wp)
            self.scene().items_ref.append(wp); self.scene().addItem(wp); self.update_position()
            self._drag_start_pos = None; self._potential_waypoint_index = -1
            wp.grabMouse(); wp.setPos(sx, sy); event.accept(); return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event): self._drag_start_pos = None; self._potential_waypoint_index = -1; super().mouseReleaseEvent(event)
    
    def mouseDoubleClickEvent(self, event):
        new_text, ok = QInputDialog.getMultiLineText(None, "エッジのテキスト編集", "線上のテキスト:", self.raw_text)
        if ok:
            self.set_text(new_text)
            if self.scene() and hasattr(self.scene(), 'main_window'):
                self.scene().main_window.push_undo_state("エッジテキスト変更")
        super().mouseDoubleClickEvent(event)

    def remove_waypoint(self, wp):
        if wp in self.waypoints:
            self.waypoints.remove(wp); self.scene().removeItem(wp)
            if wp in self.scene().items_ref: self.scene().items_ref.remove(wp)
            self.update_position(); self.scene().main_window.push_undo_state("ウェイポイント削除")
            
    def check_waypoint_straightness(self, wp):
        if wp not in self.waypoints or self.routing == "orthogonal": return
        idx = self.waypoints.index(wp)
        p1 = self.source_node.scenePos() if idx == 0 else self.waypoints[idx - 1].scenePos()
        p2 = self.target_node.scenePos() if idx == len(self.waypoints) - 1 else self.waypoints[idx + 1].scenePos()
        line = QLineF(p1, p2); length = line.length()
        if length == 0: self.remove_waypoint(wp); return
        dist = abs((p2.x()-p1.x())*(p1.y()-wp.scenePos().y()) - (p1.x()-wp.scenePos().x())*(p2.y()-p1.y())) / length
        dot = (wp.scenePos().x()-p1.x())*(p2.x()-p1.x()) + (wp.scenePos().y()-p1.y())*(p2.y()-p1.y())
        if dist < 15.0 and 0 <= dot <= length ** 2: self.remove_waypoint(wp)
        self.scene().main_window.push_undo_state("線の変形")


class FlowchartScene(QGraphicsScene):
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main_window = main_window
        self.source_node = None
        self.items_ref = [] 
        self.preview_node = None
        self.preview_items = []

    def hide_preview_node(self):
        try:
            if self.preview_node: self.preview_node.hide()
            for pi in self.preview_items: pi.hide()
        except RuntimeError:
            self.preview_node = None
            self.preview_items = []

    def update_preview_node(self, pos=None, tool=None):
        if tool is None: tool = self.main_window.current_tool

        if tool not in ["process", "decision", "data", "terminal"]:
            try:
                if self.preview_node:
                    self.removeItem(self.preview_node)
                    self.preview_node = None
            except RuntimeError: self.preview_node = None

        if tool != "paste":
            try:
                if self.preview_items:
                    for pi in self.preview_items: 
                        if pi.scene() == self: self.removeItem(pi)
                    self.preview_items = []
            except RuntimeError: self.preview_items = []

        if tool in ["process", "decision", "data", "terminal"]:
            try:
                if self.preview_node and self.preview_node.node_type != tool:
                    self.removeItem(self.preview_node); self.preview_node = None
            except RuntimeError: self.preview_node = None

            if not self.preview_node:
                self.preview_node = NodeItem(0, 0, text="Node", node_type=tool)
                self.preview_node.setOpacity(0.5); self.preview_node.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
                self.preview_node.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
                self.preview_node.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False); self.preview_node.setZValue(1000)
                self.addItem(self.preview_node)
            try:
                self.preview_node.show()
                if pos is not None:
                    sx, sy = round(pos.x() / GRID_SIZE) * GRID_SIZE, round(pos.y() / GRID_SIZE) * GRID_SIZE
                    self.preview_node.setPos(sx, sy)
            except RuntimeError: self.preview_node = None

        elif tool == "paste" and self.main_window.clipboard_data:
            try:
                if not self.preview_items:
                    id_map = {}
                    for n in self.main_window.clipboard_data.get("nodes", []):
                        node = NodeItem(n["x"], n["y"], n["text"], n["type"], str(uuid.uuid4()), n.get("bg_color", "#E1F5FE"), n.get("text_color", "#000000"))
                        node.setOpacity(0.5); node.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
                        node.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
                        node.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False); node.setZValue(1000)
                        self.addItem(node); self.preview_items.append(node); id_map[n["id"]] = node

                    for e in self.main_window.clipboard_data.get("edges", []):
                        src, tgt = id_map.get(e["source"]), id_map.get(e["target"])
                        if src and tgt:
                            edge = EdgeItem(src, tgt, e.get("label", ""), e.get("width", 2), e.get("style", "solid"), e.get("routing", "straight"))
                            edge.setOpacity(0.5); edge.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
                            edge.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False); edge.setZValue(1000)
                            if e.get("text_offset"): edge.text_item.manual_offset = QPointF(e.get("text_offset")["x"], e.get("text_offset")["y"])
                            for w in e.get("waypoints", []):
                                wp = WaypointItem(w["x"], w["y"], edge)
                                wp.setOpacity(0.5); wp.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
                                wp.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False); wp.setZValue(1000)
                                edge.waypoints.append(wp); self.addItem(wp); self.preview_items.append(wp)
                            src.add_edge(edge); tgt.add_edge(edge)
                            self.addItem(edge); self.preview_items.append(edge); edge.update_position()
                
                if pos is not None and self.main_window.clipboard_base_pos:
                    sx, sy = round(pos.x() / GRID_SIZE) * GRID_SIZE, round(pos.y() / GRID_SIZE) * GRID_SIZE
                    dx, dy = sx - self.main_window.clipboard_base_pos.x(), sy - self.main_window.clipboard_base_pos.y()
                    
                    for item in self.preview_items:
                        item.show()
                        if isinstance(item, NodeItem) or isinstance(item, WaypointItem): 
                            item.setPos(item.orig_x + dx, item.orig_y + dy)
                            
                    for item in self.preview_items:
                        if isinstance(item, EdgeItem):
                            item.update_position()
                            
            except RuntimeError: self.preview_items = []


    def drawBackground(self, painter, rect):
        super().drawBackground(painter, rect)
        painter.setPen(QPen(QColor(200, 200, 200) if self.main_window.is_light_theme else QColor(60, 60, 60), 1, Qt.PenStyle.SolidLine))
        left, top = int(rect.left()) - (int(rect.left()) % GRID_SIZE), int(rect.top()) - (int(rect.top()) % GRID_SIZE)
        lines = [QLineF(x, rect.top(), x, rect.bottom()) for x in range(left, int(rect.right()), GRID_SIZE)]
        lines.extend([QLineF(rect.left(), y, rect.right(), y) for y in range(top, int(rect.bottom()), GRID_SIZE)])
        painter.drawLines(lines)

    def mousePressEvent(self, event):
        self.main_window.is_moving = False
        tool = self.main_window.current_tool
        
        if tool in ["process", "decision", "data", "terminal"] and event.button() == Qt.MouseButton.LeftButton:
            sx, sy = round(event.scenePos().x()/GRID_SIZE)*GRID_SIZE, round(event.scenePos().y()/GRID_SIZE)*GRID_SIZE
            node = NodeItem(sx, sy, text="Node", node_type=tool)
            self.items_ref.append(node); self.addItem(node); self.main_window.push_undo_state(f"ノード追加 ({tool})")
            return

        if tool == "connect" and event.button() == Qt.MouseButton.LeftButton:
            item = self.itemAt(event.scenePos(), QTransform())
            while item and not isinstance(item, NodeItem): item = item.parentItem()
            if isinstance(item, NodeItem):
                if self.source_node is None:
                    self.source_node = item; self.source_node.set_highlight(True)
                    self.main_window.statusBar().showMessage("エッジ接続モード: 2つ目のノードをクリック")
                elif item != self.source_node:
                    edge = EdgeItem(self.source_node, item, routing=self.main_window.cb_routing.currentData())
                    self.source_node.add_edge(edge); item.add_edge(edge)
                    self.items_ref.append(edge); self.addItem(edge)
                    self.source_node.set_highlight(False); self.source_node = None
                    self.main_window.push_undo_state("エッジ接続")
                    self.main_window.statusBar().showMessage("エッジ接続モード: 次の1つ目のノードをクリック")
                return
            else:
                if self.source_node: self.source_node.set_highlight(False); self.source_node = None
                return

        if tool == "paste" and event.button() == Qt.MouseButton.LeftButton:
            if self.main_window.clipboard_data and self.main_window.clipboard_base_pos:
                sx, sy = round(event.scenePos().x() / GRID_SIZE) * GRID_SIZE, round(event.scenePos().y() / GRID_SIZE) * GRID_SIZE
                dx, dy = sx - self.main_window.clipboard_base_pos.x(), sy - self.main_window.clipboard_base_pos.y()
                self.main_window.scene.clearSelection()
                try:
                    for pi in self.preview_items: 
                        if pi.scene() == self: self.removeItem(pi)
                except RuntimeError: pass
                self.preview_items = []
                self.main_window.load_scene_json(self.main_window.clipboard_data, offset_x=dx, offset_y=dy, clear_scene=False, generate_new_ids=True)
                self.main_window.push_undo_state("貼り付け")
                self.main_window.set_tool("select")
                self.main_window.btn_select.setChecked(True)
            return

        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self.selectedItems(): self.main_window.is_moving = True
        tool = self.main_window.current_tool
        if tool in ["process", "decision", "data", "terminal", "paste"]: self.update_preview_node(event.scenePos(), tool)
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        super().mouseReleaseEvent(event)
        if getattr(self.main_window, 'is_moving', False):
            self.main_window.push_undo_state("移動"); self.main_window.is_moving = False

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace): self.main_window.delete_selected_items()
        super().keyPressEvent(event)


def clip_line_to_node(p_start: QPointF, p_end: QPointF, node: NodeItem) -> QPointF:
    line = QLineF(p_start, p_end); polygon = node.mapToScene(node.path().toFillPolygon())
    best_p, min_dist = p_start, float('inf')
    for i in range(polygon.count()):
        p_a, p_b = polygon.at(i), polygon.at((i + 1) % polygon.count())
        intersect_type, ip = line.intersects(QLineF(p_a, p_b))
        if intersect_type == QLineF.IntersectionType.BoundedIntersection:
            dist = QLineF(p_start, ip).length()
            if dist < min_dist: min_dist = dist; best_p = ip
    return best_p


class CustomPrintPreviewDialog(QDialog):
    def __init__(self, main_window, has_selection=False):
        super().__init__(main_window); self.main_window = main_window
        self.setWindowTitle("印刷プレビューと設定"); self.resize(1100, 750)
        self.printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        
        settings_layout = QVBoxLayout()
        printer_group = QGroupBox("プリンタ"); pr_layout = QVBoxLayout()
        self.printer_combo = QComboBox()
        for p in QPrinterInfo.availablePrinters(): self.printer_combo.addItem(p.printerName(), p)
        idx = self.printer_combo.findText(QPrinterInfo.defaultPrinter().printerName())
        if idx >= 0: self.printer_combo.setCurrentIndex(idx)
        pr_layout.addWidget(self.printer_combo); printer_group.setLayout(pr_layout); settings_layout.addWidget(printer_group)
        
        paper_group = QGroupBox("用紙設定"); pp_layout = QVBoxLayout()
        h_size = QHBoxLayout(); h_size.addWidget(QLabel("サイズ:"))
        self.paper_size_combo = QComboBox()
        for name, sz in [("A4", QPageSize.PageSizeId.A4), ("A3", QPageSize.PageSizeId.A3), ("B5", QPageSize.PageSizeId.B5), ("B4", QPageSize.PageSizeId.B4), ("Letter", QPageSize.PageSizeId.Letter)]:
            self.paper_size_combo.addItem(name, sz)
        h_size.addWidget(self.paper_size_combo); pp_layout.addLayout(h_size)
        
        h_ori = QHBoxLayout(); h_ori.addWidget(QLabel("向き:"))
        self.ori_portrait = QRadioButton("縦"); self.ori_landscape = QRadioButton("横")
        self.ori_portrait.setChecked(True); h_ori.addWidget(self.ori_portrait); h_ori.addWidget(self.ori_landscape); pp_layout.addLayout(h_ori)
        
        h_margin = QHBoxLayout(); h_margin.addWidget(QLabel("余白(mm):"))
        self.margin_spin = QDoubleSpinBox(); self.margin_spin.setRange(0, 100); self.margin_spin.setValue(10.0)
        h_margin.addWidget(self.margin_spin); pp_layout.addLayout(h_margin); paper_group.setLayout(pp_layout); settings_layout.addWidget(paper_group)
        
        range_group = QGroupBox("印刷範囲"); range_layout = QVBoxLayout()
        self.radio_all = QRadioButton("図面全体"); self.radio_view = QRadioButton("現在の表示範囲"); self.radio_sel = QRadioButton("選択したアイテム")
        self.radio_all.setChecked(True); self.radio_sel.setEnabled(has_selection)
        for r in [self.radio_all, self.radio_view, self.radio_sel]: range_layout.addWidget(r)
        range_group.setLayout(range_layout); settings_layout.addWidget(range_group)
        
        scale_group = QGroupBox("スケール設定"); sc_layout = QVBoxLayout()
        self.radio_auto = QRadioButton("自動調整"); self.radio_custom = QRadioButton("倍率指定(%)"); self.radio_auto.setChecked(True)
        h_scale = QHBoxLayout(); self.spin_scale = QDoubleSpinBox(); self.spin_scale.setRange(10, 1000); self.spin_scale.setValue(100); self.spin_scale.setEnabled(False)
        h_scale.addWidget(self.radio_custom); h_scale.addWidget(self.spin_scale); sc_layout.addWidget(self.radio_auto); sc_layout.addLayout(h_scale); scale_group.setLayout(sc_layout); settings_layout.addWidget(scale_group)
        
        btn_layout = QVBoxLayout(); self.btn_print = QPushButton("🖨️ 印刷を実行"); self.btn_print.setStyleSheet("font-weight: bold; padding: 10px;")
        self.btn_cancel = QPushButton("キャンセル"); btn_layout.addSpacing(20); btn_layout.addWidget(self.btn_print); btn_layout.addWidget(self.btn_cancel); settings_layout.addLayout(btn_layout)
        settings_layout.addStretch()
        
        self.preview_widget = QPrintPreviewWidget(self.printer); self.preview_widget.paintRequested.connect(self.handle_paint_request)
        main_layout = QHBoxLayout(self); left_panel = QWidget(); left_panel.setLayout(settings_layout); left_panel.setFixedWidth(280)
        main_layout.addWidget(left_panel); main_layout.addWidget(self.preview_widget, stretch=1)
        
        self.radio_custom.toggled.connect(self.spin_scale.setEnabled)
        for w in [self.printer_combo, self.paper_size_combo, self.ori_portrait, self.ori_landscape, self.margin_spin, self.radio_all, self.radio_view, self.radio_sel, self.radio_auto, self.spin_scale]:
            if isinstance(w, QComboBox): w.currentIndexChanged.connect(self.update_preview)
            elif isinstance(w, QRadioButton): w.toggled.connect(self.update_preview)
            elif isinstance(w, QDoubleSpinBox): w.valueChanged.connect(self.update_preview)
        self.btn_print.clicked.connect(self.do_print); self.btn_cancel.clicked.connect(self.reject)
        self.update_preview()

    def update_printer_settings(self):
        if self.printer_combo.currentData(): self.printer.setPrinterName(self.printer_combo.currentData().printerName())
        self.printer.setPageSize(QPageSize(self.paper_size_combo.currentData()))
        self.printer.setPageOrientation(QPageLayout.Orientation.Portrait if self.ori_portrait.isChecked() else QPageLayout.Orientation.Landscape)
        m = self.margin_spin.value(); self.printer.setPageMargins(QMarginsF(m, m, m, m), QPageLayout.Unit.Millimeter)

    def update_preview(self): self.update_printer_settings(); self.preview_widget.updatePreview()

    def do_print(self):
        self.update_printer_settings(); dialog = QPrintDialog(self.printer, self)
        if dialog.exec() == QPrintDialog.DialogCode.Accepted: self.handle_paint_request(self.printer); self.accept()

    def handle_paint_request(self, printer):
        if self.radio_all.isChecked(): print_rect = self.main_window.scene.itemsBoundingRect(); sel_only = False
        elif self.radio_view.isChecked(): print_rect = self.main_window.view.mapToScene(self.main_window.view.viewport().rect()).boundingRect(); sel_only = False
        else:
            rect = QRectF()
            for si in self.main_window.scene.selectedItems(): rect = rect.united(si.sceneBoundingRect())
            print_rect = rect; sel_only = True

        if print_rect.isEmpty(): return
        sel_items = self.main_window.scene.selectedItems(); self.main_window.scene.clearSelection()
        rect = QRectF(print_rect).adjusted(-5, -5, 5, 5); painter = QPainter(printer); painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        page_rect = printer.pageRect(QPrinter.Unit.DevicePixel); hidden_items = []
        
        def should_keep(it):
            if it in sel_items: return True
            for c in it.childItems():
                if should_keep(c): return True
            return False

        for item in self.main_window.scene.items():
            if not item.isVisible(): continue
            
            is_preview = False
            try:
                if item == getattr(self.main_window.scene, 'preview_node', None) or item in getattr(self.main_window.scene, 'preview_items', []):
                    is_preview = True
            except RuntimeError: pass

            if isinstance(item, WaypointItem) or is_preview or (sel_only and item.parentItem() is None and not should_keep(item)):
                item.hide(); hidden_items.append(item)

        if self.radio_auto.isChecked(): self.main_window.scene.render(painter, QRectF(page_rect), rect, Qt.AspectRatioMode.KeepAspectRatio)
        else:
            sc = (self.spin_scale.value() / 100.0) * (printer.resolution() / self.main_window.logicalDpiX())
            sw, sh = rect.width() * sc, rect.height() * sc
            tx, ty = page_rect.left() + max(0, (page_rect.width()-sw)/2.0), page_rect.top() + max(0, (page_rect.height()-sh)/2.0)
            self.main_window.scene.render(painter, QRectF(tx, ty, sw, sh), rect, Qt.AspectRatioMode.KeepAspectRatio)
        
        for item in hidden_items: item.show()
        painter.end()
        for item in sel_items: item.setSelected(True)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_filepath = None
        self.current_tool = "select"
        self.clipboard_data = None
        self.clipboard_base_pos = None
        self.is_light_theme = True
        
        self.undo_stack = QUndoStack(self)
        self.last_state = {"nodes": [], "edges": [], "groups": []}

        self.scene = FlowchartScene(self)
        self.scene.setSceneRect(-2000, -2000, 4000, 4000)

        self.view = FlowchartView(self.scene)
        self.view.centerOn(0, 0)
        self.setCentralWidget(self.view)

        self.icon_actions = [] 
        self.init_menu()
        self.init_toolbars()
        self.apply_theme() 
        
        self.update_window_title()
        self.statusBar().showMessage("準備完了: 範囲選択や複数選択（Ctrlキー+クリック）が可能です")

    def create_icon_action(self, icon_name, text, slot=None, shortcut=None, checkable=False):
        act = QAction(text, self)
        if shortcut: act.setShortcut(shortcut)
        if checkable: act.setCheckable(True)
        if slot: act.triggered.connect(slot)
        self.icon_actions.append((act, icon_name))
        return act

    def apply_theme(self):
        theme = "light" if self.is_light_theme else "dark"
        app = QApplication.instance()
        if app:
            try:
                if hasattr(qdarktheme, 'setup_theme'):
                    qdarktheme.setup_theme(theme)
                else:
                    app.setStyleSheet(qdarktheme.load_stylesheet(theme))
            except Exception as e:
                print(f"Theme setup failed: {e}")
        
        text_color = "#212529" if self.is_light_theme else "#F8F9FA"
        bg_hover = "#E2E6EA" if self.is_light_theme else "#495057"
        
        self.setStyleSheet(f"""
            QToolBar {{ spacing: 6px; padding: 4px; border: none; }} 
            QToolButton {{ font-size: 13px; padding: 6px 10px; border-radius: 4px; color: {text_color}; }} 
            QToolButton:hover {{ background: {bg_hover}; }}
            QToolButton:checked {{ background: #3B82F6; color: white; font-weight: bold; }} 
        """)

        icon_color = '#212529' if self.is_light_theme else '#F8F9FA'
        for act, icon_name in self.icon_actions:
            if icon_name:
                act.setIcon(qta.icon(icon_name, color=icon_color))

        if self.is_light_theme:
            self.scene.setBackgroundBrush(QBrush(QColor(255, 255, 255)))
        else:
            self.scene.setBackgroundBrush(QBrush(QColor(30, 30, 30)))
        self.scene.update()

    def toggle_theme(self):
        self.is_light_theme = not self.is_light_theme
        self.apply_theme()

    def get_scene_json(self, selected_only=False):
        data = {"nodes": [], "edges": [], "groups": []}
        items_raw = self.scene.selectedItems() if selected_only else self.scene.items()
        
        items = set()
        for it in items_raw:
            items.add(it)
            if type(it) == QGraphicsItemGroup:
                for child in it.childItems():
                    items.add(child)
        items = list(items)

        valid_node_ids = set()
        
        for item in items:
            if isinstance(item, NodeItem) and getattr(self.scene, 'preview_node', None) != item and item not in getattr(self.scene, 'preview_items', []):
                data["nodes"].append({"id": item.node_id, "type": item.node_type, "x": item.scenePos().x(), "y": item.scenePos().y(), "text": item.text_item.toPlainText(), "bg_color": item.bg_color.name(), "text_color": item.text_color.name()})
                valid_node_ids.add(item.node_id)
                
        for item in items:
            if isinstance(item, EdgeItem) and item not in getattr(self.scene, 'preview_items', []):
                if selected_only and (item.source_node.node_id not in valid_node_ids or item.target_node.node_id not in valid_node_ids): continue
                offset = {"x": item.text_item.manual_offset.x(), "y": item.text_item.manual_offset.y()} if item.text_item.manual_offset else None
                data["edges"].append({"source": item.source_node.node_id, "target": item.target_node.node_id, "label": item.raw_text, "width": item.line_width, "style": item.line_style, "routing": item.routing, "waypoints": [{"x": wp.scenePos().x(), "y": wp.scenePos().y()} for wp in item.waypoints], "text_offset": offset})
                
        for item in items:
            if type(item) == QGraphicsItemGroup:
                c_ids = [c.node_id for c in item.childItems() if hasattr(c, 'node_id')]
                if c_ids: data["groups"].append(c_ids)
        return data

    def load_scene_json(self, data, offset_x=0, offset_y=0, clear_scene=True, generate_new_ids=False, is_undo_redo=False):
        if clear_scene: 
            self.scene.clear(); self.scene.items_ref.clear()
            self.scene.preview_node = None; self.scene.preview_items = []; self.scene.source_node = None
            
        id_map = {}
        for n in data.get("nodes", []):
            new_id = str(uuid.uuid4()) if generate_new_ids else n.get("id", str(uuid.uuid4()))
            node = NodeItem(n["x"]+offset_x, n["y"]+offset_y, n["text"], n["type"], new_id, n.get("bg_color", "#E1F5FE"), n.get("text_color", "#000000"))
            self.scene.items_ref.append(node); self.scene.addItem(node); id_map[n.get("id")] = node
            if not clear_scene: node.setSelected(True)
            
        for e in data.get("edges", []):
            src, tgt = id_map.get(e["source"]), id_map.get(e["target"])
            if src and tgt:
                edge = EdgeItem(src, tgt, e.get("label", ""), e.get("width", 2), e.get("style", "solid"), e.get("routing", "straight"))
                if e.get("text_offset"): edge.text_item.manual_offset = QPointF(e.get("text_offset")["x"], e.get("text_offset")["y"])
                for w in e.get("waypoints", []):
                    wp = WaypointItem(w["x"]+offset_x, w["y"]+offset_y, edge); edge.waypoints.append(wp); self.scene.items_ref.append(wp); self.scene.addItem(wp)
                src.add_edge(edge); tgt.add_edge(edge); self.scene.items_ref.append(edge); self.scene.addItem(edge); edge.update_position()
                if not clear_scene: edge.setSelected(True)
                
        for g_cids in data.get("groups", []):
            g_items = [id_map[cid] for cid in g_cids if cid in id_map]
            if g_items:
                group = self.scene.createItemGroup(g_items)
                group.setFlags(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable | QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
                self.scene.items_ref.append(group); getattr(group, 'setSelected', lambda x: None)(True if not clear_scene else False)
                
        if not is_undo_redo: self.last_state = self.get_scene_json()

    def push_undo_state(self, description):
        new_state = self.get_scene_json()
        if self.last_state != new_state:
            self.undo_stack.push(SceneStateCommand(self, self.last_state, new_state, description))
            self.last_state = new_state

    def update_window_title(self):
        base = "FlowchartCreationMiya v1.2.0"
        self.setWindowTitle(f"{os.path.basename(self.current_filepath)} - {base}" if self.current_filepath else base)

    def init_menu(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu("ファイル(&F)")
        for text, func, sc in [("上書き保存(&S)", self.save_file, "Ctrl+S"), ("名前を付けて保存(&A)...", self.save_file_as, "Ctrl+Shift+S"), ("読込(&O)", self.load_json, "Ctrl+O")]:
            act = QAction(text, self); act.setShortcut(sc); act.triggered.connect(func); file_menu.addAction(act)
        file_menu.addSeparator()
        act_excel = self.create_icon_action('fa5s.file-excel', "仕様書(Excel)生成...", self.generate_excel)
        file_menu.addAction(act_excel)
        file_menu.addSeparator()
        file_menu.addAction("Draw.ioインポート(.xml)...", self.import_drawio)
        file_menu.addAction("Draw.ioエクスポート(.xml)...", self.export_drawio)
        file_menu.addSeparator()
        file_menu.addAction("画像エクスポート(&E)...", self.export_file); file_menu.addAction("Jw_cadへコピー(&C)", self.copy_to_jwcad)
        file_menu.addSeparator(); file_menu.addAction("終了(&X)", self.close)
        
        edit_menu = menubar.addMenu("編集(&E)")
        self.act_undo = self.undo_stack.createUndoAction(self, "元に戻す(&U)")
        self.act_undo.setShortcut("Ctrl+Z")
        self.icon_actions.append((self.act_undo, 'fa5s.undo'))
        edit_menu.addAction(self.act_undo)

        self.act_redo = self.undo_stack.createRedoAction(self, "やり直し(&R)")
        self.act_redo.setShortcut("Ctrl+Y")
        self.icon_actions.append((self.act_redo, 'fa5s.redo'))
        edit_menu.addAction(self.act_redo)
        edit_menu.addSeparator()
        
        act_copy = self.create_icon_action('fa5s.copy', "コピー(&C)", self.copy_items, shortcut="Ctrl+C")
        act_paste = self.create_icon_action('fa5s.paste', "貼り付け(&V)", self.paste_items, shortcut="Ctrl+V")
        act_del = self.create_icon_action('fa5s.trash-alt', "削除(&D)", self.delete_selected_items, shortcut="Del")
        edit_menu.addAction(act_copy); edit_menu.addAction(act_paste); edit_menu.addAction(act_del)
        edit_menu.addSeparator()
        
        act_grp = QAction("グループ化(&G)", self); act_grp.setShortcut("Ctrl+G"); act_grp.triggered.connect(self.group_selected); edit_menu.addAction(act_grp)
        act_ungrp = QAction("グループ解除(&U)", self); act_ungrp.setShortcut("Ctrl+Shift+G"); act_ungrp.triggered.connect(self.ungroup_selected); edit_menu.addAction(act_ungrp)
        
        arr_menu = menubar.addMenu("配置(&A)")
        act_layout = self.create_icon_action('fa5s.sitemap', "★自動階層レイアウト", self.auto_layout_networkx)
        arr_menu.addAction(act_layout)
        arr_menu.addSeparator()
        for txt, mode in [("左揃え", "left"), ("左右中央揃え", "center_x"), ("右揃え", "right"), ("上揃え", "top"), ("上下中央揃え", "center_y"), ("下揃え", "bottom"), ("水平等間隔", "dist_h"), ("垂直等間隔", "dist_v")]:
            act = QAction(txt, self); act.triggered.connect(lambda chk, m=mode: self.align_items(m)); arr_menu.addAction(act)

        view_menu = menubar.addMenu("表示(&V)")
        act_theme = self.create_icon_action('fa5s.adjust', "テーマ切り替え (Light/Dark)", self.toggle_theme)
        view_menu.addAction(act_theme)
        
        help_menu = menubar.addMenu("ヘルプ(&H)")
        help_menu.addAction("使い方(&U)", self.show_usage); help_menu.addAction("バージョン情報(&A)", self.show_about)

    def show_usage(self):
        msg = ("【操作説明】\n"
               "・図形の追従プレビュー: 追加モード時やコピペ時、カーソルにゴーストが追従します。\n"
               "・Undo/Redo: Ctrl+Z / Ctrl+Y\n"
               "・コピー＆ペースト: Ctrl+Cでコピーし、キャンバスをクリックして配置\n"
               "・グループ化: Ctrl+G / 解除: Ctrl+Shift+G\n"
               "・整列 / 自動レイアウト: 複数選択して上部メニューの「配置」から実行\n"
               "・線のスタイル: エッジを選択して「書式ツールバー」で太さや直角配線を変更\n\n"
               "・Jw_cad連携 / 仕様書生成 / Draw.io互換 は「ファイル」メニューから実行可能です。")
        QMessageBox.information(self, "使い方", msg)

    def show_about(self): 
        QMessageBox.about(self, "情報", "FlowchartCreationMiya v1.2.0\nPython & PyQt6 製フローチャート作成ツール")

    def init_toolbars(self):
        tb_main = QToolBar("メインツール"); self.addToolBar(tb_main)
        self.action_group = QActionGroup(self)
        self.btn_select = self.create_icon_action('fa5s.mouse-pointer', "選択", lambda: self.set_tool("select"), checkable=True)
        self.btn_select.setChecked(True)
        tb_main.addAction(self.btn_select)
        self.action_group.addAction(self.btn_select)
        tb_main.addSeparator()
        
        for icon_name, text, key in [("fa5s.square", "処理", "process"), ("fa5s.code-branch", "分岐", "decision"), ("fa5s.layer-group", "データ", "data"), ("fa5s.capsules", "端子", "terminal")]:
            act = self.create_icon_action(icon_name, text, lambda chk, k=key: self.set_tool(k), checkable=True)
            tb_main.addAction(act); self.action_group.addAction(act)
        
        tb_main.addSeparator()
        act_conn = self.create_icon_action('fa5s.link', "接続", lambda: self.set_tool("connect"), checkable=True)
        tb_main.addAction(act_conn); self.action_group.addAction(act_conn)

        tb_edit = QToolBar("編集"); self.addToolBar(tb_edit)
        tb_edit.addAction(self.act_undo)
        tb_edit.addAction(self.act_redo)
        tb_edit.addSeparator()
        for act, _ in self.icon_actions:
            if act.text() in ["コピー(&C)", "貼り付け(&V)", "削除(&D)"]:
                tb_edit.addAction(act)
        
        self.addToolBarBreak()

        tb_style = QToolBar("書式"); self.addToolBar(tb_style)
        act_bg = self.create_icon_action('fa5s.fill-drip', "背景", self.change_bg_color)
        act_fg = self.create_icon_action('fa5s.font', "文字色", self.change_text_color)
        tb_style.addAction(act_bg)
        tb_style.addAction(act_fg)
        tb_style.addSeparator()
        tb_style.addWidget(QLabel("線幅:")); self.cb_width = QComboBox(); self.cb_width.addItems(["1", "2", "3", "4", "5"]); self.cb_width.setCurrentText("2"); self.cb_width.currentTextChanged.connect(self.change_edge_style); tb_style.addWidget(self.cb_width)
        tb_style.addWidget(QLabel("線種:")); self.cb_style = QComboBox(); self.cb_style.addItems(["実線(solid)", "破線(dash)", "点線(dot)"]); self.cb_style.currentTextChanged.connect(self.change_edge_style); tb_style.addWidget(self.cb_style)
        tb_style.addWidget(QLabel("ルート:")); self.cb_routing = QComboBox(); self.cb_routing.addItems(["直線", "直角(Orthogonal)"]); self.cb_routing.setItemData(0, "straight"); self.cb_routing.setItemData(1, "orthogonal"); self.cb_routing.currentIndexChanged.connect(self.change_edge_style); tb_style.addWidget(self.cb_routing)

    def set_tool(self, tool_name):
        self.current_tool = tool_name
        if self.scene.source_node: self.scene.source_node.set_highlight(False); self.scene.source_node = None
        
        if tool_name == "select": 
            self.view.setDragMode(QGraphicsView.DragMode.RubberBandDrag); self.view.setCursor(Qt.CursorShape.ArrowCursor); self.statusBar().showMessage("準備完了")
            self.scene.update_preview_node(None, tool_name)
        elif tool_name == "connect": 
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag); self.view.setCursor(Qt.CursorShape.CrossCursor); self.statusBar().showMessage("エッジ接続: 1つ目のノードをクリック")
            self.scene.update_preview_node(None, tool_name)
        elif tool_name == "paste":
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag); self.view.setCursor(Qt.CursorShape.CrossCursor); self.statusBar().showMessage("ペーストモード")
            g_pos = QCursor.pos(); v_pos = self.view.mapFromGlobal(g_pos)
            if self.view.rect().contains(v_pos): self.scene.update_preview_node(self.view.mapToScene(v_pos), tool_name)
        else: 
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag); self.view.setCursor(Qt.CursorShape.CrossCursor); self.statusBar().showMessage(f"配置モード: {tool_name}")
            g_pos = QCursor.pos(); v_pos = self.view.mapFromGlobal(g_pos)
            if self.view.rect().contains(v_pos): self.scene.update_preview_node(self.view.mapToScene(v_pos), tool_name)

    def auto_layout_networkx(self):
        data = self.get_scene_json()
        if not data["nodes"]: return

        G = nx.DiGraph()
        for n in data["nodes"]: G.add_node(n["id"])
        for e in data["edges"]: G.add_edge(e["source"], e["target"])

        try:
            for layer, nodes in enumerate(nx.topological_generations(G) if nx.is_directed_acyclic_graph(G) else [list(G.nodes)]):
                for node in nodes:
                    G.nodes[node]["layer"] = layer
            pos = nx.multipartite_layout(G, subset_key="layer", align="horizontal")
            
            x_scale, y_scale = 200, 150
            for i in self.scene.items():
                if isinstance(i, NodeItem) and i.node_id in pos:
                    p = pos[i.node_id]
                    sx, sy = round(p[0]*x_scale/GRID_SIZE)*GRID_SIZE, round(p[1]*y_scale/GRID_SIZE)*GRID_SIZE
                    i.setPos(sx, sy)
            self.push_undo_state("自動レイアウト")
            QMessageBox.information(self, "完了", "自動レイアウトが完了しました。")
        except Exception as e:
            QMessageBox.warning(self, "エラー", f"自動レイアウトに失敗しました。\n{e}")

    def generate_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "仕様書(Excel)生成", "", "Excel Files (*.xlsx)")
        if not path: return
        
        data = self.get_scene_json()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Flowchart Specifications"
        
        headers = ["Step", "Node ID", "Type", "Text", "Next Steps"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True)

        id_to_node = {n["id"]: n for n in data["nodes"]}
        edges_from = {}
        for e in data["edges"]:
            edges_from.setdefault(e["source"], []).append(e["target"])

        row = 2
        for i, n in enumerate(data["nodes"], 1):
            next_ids = edges_from.get(n["id"], [])
            next_texts = [id_to_node[nid]["text"].replace('\n', ' ') for nid in next_ids if nid in id_to_node]
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=n["id"])
            ws.cell(row=row, column=3, value=n["type"])
            ws.cell(row=row, column=4, value=n["text"])
            ws.cell(row=row, column=5, value=", ".join(next_texts))
            row += 1

        try:
            wb.save(path)
            QMessageBox.information(self, "完了", f"仕様書を生成しました:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"保存に失敗しました:\n{e}")

    def import_drawio(self):
        path, _ = QFileDialog.getOpenFileName(self, "Draw.io インポート", "", "XML Files (*.xml *.drawio)")
        if not path: return
        try:
            tree = ET.parse(path)
            root = tree.getroot()
            
            self.scene.clear(); self.scene.items_ref.clear()
            id_map = {}
            for cell in root.iter('mxCell'):
                cid = cell.get('id')
                geom = cell.find('mxGeometry')
                if geom is None: continue
                
                if cell.get('edge') == '1':
                    src, tgt = cell.get('source'), cell.get('target')
                    if src in id_map and tgt in id_map:
                        edge = EdgeItem(id_map[src], id_map[tgt], cell.get('value', ''), 2, "solid", "orthogonal")
                        self.scene.items_ref.append(edge); self.scene.addItem(edge); edge.update_position()
                elif cell.get('vertex') == '1':
                    x, y = float(geom.get('x', 0)), float(geom.get('y', 0))
                    txt = cell.get('value', '').replace('&lt;br&gt;', '\n').replace('<br>', '\n')
                    style = cell.get('style', '')
                    ntype = "process"
                    if "rhombus" in style: ntype = "decision"
                    elif "ellipse" in style: ntype = "terminal"
                    node = NodeItem(x, y, txt, ntype, cid)
                    self.scene.items_ref.append(node); self.scene.addItem(node); id_map[cid] = node

            self.last_state = self.get_scene_json()
            QMessageBox.information(self, "完了", "Draw.ioファイルのインポートが完了しました。")
        except Exception as e:
            QMessageBox.warning(self, "エラー", f"読み込みに失敗しました:\n{e}")

    def export_drawio(self):
        path, _ = QFileDialog.getSaveFileName(self, "Draw.io エクスポート", "", "XML Files (*.xml)")
        if not path: return
        data = self.get_scene_json()
        
        mxfile = ET.Element('mxfile')
        diagram = ET.SubElement(mxfile, 'diagram', id=str(uuid.uuid4()), name="Page-1")
        mxGraphModel = ET.SubElement(diagram, 'mxGraphModel', dx="1000", dy="1000", grid="1", gridSize="20", guides="1", tooltips="1", connect="1", arrows="1", fold="1", page="1", pageScale="1", pageWidth="827", pageHeight="1169", math="0", shadow="0")
        root = ET.SubElement(mxGraphModel, 'root')
        ET.SubElement(root, 'mxCell', id="0")
        ET.SubElement(root, 'mxCell', id="1", parent="0")

        for n in data["nodes"]:
            style = "rounded=0;whiteSpace=wrap;html=1;"
            if n["type"] == "decision": style = "rhombus;whiteSpace=wrap;html=1;"
            elif n["type"] == "terminal": style = "ellipse;whiteSpace=wrap;html=1;"
            elif n["type"] == "data": style = "shape=parallelogram;perimeter=parallelogramPerimeter;whiteSpace=wrap;html=1;fixedSize=1;"
            
            cell = ET.SubElement(root, 'mxCell', id=n["id"], value=n["text"].replace('\n', '<br>'), style=style, vertex="1", parent="1")
            ET.SubElement(cell, 'mxGeometry', x=str(n["x"]-50), y=str(n["y"]-25), width="100", height="50", **{'as': 'geometry'})

        for i, e in enumerate(data["edges"]):
            style = "edgeStyle=orthogonalEdgeStyle;rounded=0;orthogonalLoop=1;jettySize=auto;html=1;" if e.get("routing")=="orthogonal" else "html=1;"
            cell = ET.SubElement(root, 'mxCell', id=f"edge_{i}", value=e.get("label", ""), style=style, edge="1", parent="1", source=e["source"], target=e["target"])
            ET.SubElement(cell, 'mxGeometry', relative="1", **{'as': 'geometry'})

        tree = ET.ElementTree(mxfile)
        try:
            tree.write(path, encoding='utf-8', xml_declaration=True)
            QMessageBox.information(self, "完了", f"Draw.io形式でエクスポートしました。\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "エラー", str(e))

    def copy_items(self):
        sel_items = self.scene.selectedItems()
        if not sel_items: return
        
        self.clipboard_data = self.get_scene_json(selected_only=True)
        nodes_data = self.clipboard_data.get("nodes", [])
        if nodes_data:
            self.clipboard_base_pos = QPointF(min(n["x"] for n in nodes_data), min(n["y"] for n in nodes_data))
            self.statusBar().showMessage("コピーしました（クリックで配置）", 3000)
            
            self.scene.clearSelection()
            self.set_tool("paste")
        else:
            self.clipboard_base_pos = QPointF(0, 0)
            self.statusBar().showMessage("コピーに失敗しました", 3000)

    def paste_items(self):
        if self.clipboard_data and self.clipboard_data.get("nodes"): 
            self.set_tool("paste")

    def group_selected(self):
        items = [i for i in self.scene.selectedItems() if isinstance(i, NodeItem) and i.parentItem() is None]
        if len(items) < 2: return
        group = self.scene.createItemGroup(items); group.setFlags(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable | QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
        self.scene.items_ref.append(group); self.push_undo_state("グループ化")

    def ungroup_selected(self):
        changed = False
        for item in self.scene.selectedItems():
            if type(item) == QGraphicsItemGroup:
                self.scene.destroyItemGroup(item)
                if item in self.scene.items_ref: self.scene.items_ref.remove(item)
                changed = True
        if changed: self.push_undo_state("グループ解除")

    def align_items(self, mode):
        nodes = [item for item in self.scene.selectedItems() if isinstance(item, NodeItem)]
        if len(nodes) < 2: return
        xs = [n.scenePos().x() for n in nodes]; ys = [n.scenePos().y() for n in nodes]
        if mode == "left": val = min(xs); [n.setPos(val, n.scenePos().y()) for n in nodes]
        elif mode == "right": val = max(xs); [n.setPos(val, n.scenePos().y()) for n in nodes]
        elif mode == "center_x": val = sum(xs)/len(xs); [n.setPos(val, n.scenePos().y()) for n in nodes]
        elif mode == "top": val = min(ys); [n.setPos(n.scenePos().x(), val) for n in nodes]
        elif mode == "bottom": val = max(ys); [n.setPos(n.scenePos().x(), val) for n in nodes]
        elif mode == "center_y": val = sum(ys)/len(ys); [n.setPos(n.scenePos().x(), val) for n in nodes]
        elif mode == "dist_h":
            nodes.sort(key=lambda n: n.scenePos().x()); span = (max(xs) - min(xs)) / (len(nodes) - 1)
            for i, n in enumerate(nodes): n.setPos(min(xs) + i*span, n.scenePos().y())
        elif mode == "dist_v":
            nodes.sort(key=lambda n: n.scenePos().y()); span = (max(ys) - min(ys)) / (len(nodes) - 1)
            for i, n in enumerate(nodes): n.setPos(n.scenePos().x(), min(ys) + i*span)
        self.push_undo_state(f"整列")

    def change_bg_color(self):
        nodes = [i for i in self.scene.selectedItems() if isinstance(i, NodeItem)]
        if not nodes: return
        c = QColorDialog.getColor(nodes[0].bg_color, self, "背景色")
        if c.isValid(): [n.set_bg_color(c) for n in nodes]; self.push_undo_state("背景色変更")

    def change_text_color(self):
        nodes = [i for i in self.scene.selectedItems() if isinstance(i, NodeItem)]
        if not nodes: return
        c = QColorDialog.getColor(nodes[0].text_color, self, "文字色")
        if c.isValid(): [n.set_text_color(c) for n in nodes]; self.push_undo_state("文字色変更")

    def change_edge_style(self):
        edges = [i for i in self.scene.selectedItems() if isinstance(i, EdgeItem)]
        if not edges: return
        w = int(self.cb_width.currentText()); s = self.cb_style.currentText().split("(")[1].replace(")","")
        r = self.cb_routing.currentData()
        for e in edges: e.line_width = w; e.line_style = s; e.routing = r; e.update_pen(); e.update_position()
        self.push_undo_state("線のスタイル変更")

    def delete_selected_items(self):
        sel = self.scene.selectedItems()
        if not sel: return
        edges, nodes, wps = set(), set(), set()
        for i in sel:
            if type(i) == QGraphicsItemGroup: self.ungroup_selected(); return 
            if isinstance(i, NodeItem): nodes.add(i); edges.update(i.edges)
            elif isinstance(i, EdgeItem): edges.add(i)
            elif isinstance(i, WaypointItem): wps.add(i)
        for wp in wps:
            if wp.edge not in edges: wp.edge.remove_waypoint(wp)
        for e in edges:
            for wp in e.waypoints:
                if wp.scene(): self.scene.removeItem(wp)
                if wp in self.scene.items_ref: self.scene.items_ref.remove(wp)
            e.waypoints.clear()
            if e in e.source_node.edges: e.source_node.edges.remove(e)
            if e in e.target_node.edges: e.target_node.edges.remove(e)
            if e.scene(): self.scene.removeItem(e)
            if e in self.scene.items_ref: self.scene.items_ref.remove(e)
        for n in nodes:
            if n.scene(): self.scene.removeItem(n)
            if n in self.scene.items_ref: self.scene.items_ref.remove(n)
        self.push_undo_state("削除")

    def zoom_in(self): self.view.scale(1.15, 1.15)
    def zoom_out(self): self.view.scale(1/1.15, 1/1.15)
    def zoom_reset(self): self.view.resetTransform()

    def save_file(self):
        if self.current_filepath:
            with open(self.current_filepath, 'w', encoding='utf-8') as f: json.dump(self.get_scene_json(), f, indent=4, ensure_ascii=False)
            self.statusBar().showMessage("上書き保存しました", 5000)
        else: self.save_file_as()

    def save_file_as(self):
        path, _ = QFileDialog.getSaveFileName(self, "名前を付けて保存", "", "JSON Files (*.json)")
        if path: self.current_filepath = path; self.save_file(); self.update_window_title()

    def load_json(self):
        path, _ = QFileDialog.getOpenFileName(self, "読込", "", "JSON Files (*.json)")
        if path:
            with open(path, 'r', encoding='utf-8') as f: data = json.load(f)
            self.load_scene_json(data); self.undo_stack.clear(); self.last_state = self.get_scene_json(); self.current_filepath = path; self.update_window_title()

    def export_file(self):
        path, _ = QFileDialog.getSaveFileName(self, "エクスポート", "", "PNG Files (*.png);;JPEG Files (*.jpeg *.jpg);;SVG Files (*.svg);;DXF Files (*.dxf)")
        if not path: return
        self.scene.clearSelection(); rect = self.scene.itemsBoundingRect().adjusted(-20, -20, 20, 20)
        hidden = []
        for i in self.scene.items():
            try:
                if (isinstance(i, WaypointItem) or i == getattr(self.scene, 'preview_node', None) or i in getattr(self.scene, 'preview_items', [])) and i.isVisible():
                    i.hide(); hidden.append(i)
            except RuntimeError: pass
        if not rect.isEmpty():
            if path.endswith(('.png', '.jpg', '.jpeg')):
                img = QImage(rect.size().toSize(), QImage.Format.Format_ARGB32); img.fill(Qt.GlobalColor.white)
                p = QPainter(img); self.scene.render(p, QRectF(img.rect()), rect); p.end(); img.save(path)
            elif path.endswith('.svg'):
                gen = QSvgGenerator(); gen.setFileName(path); gen.setSize(rect.size().toSize()); gen.setViewBox(rect)
                p = QPainter(gen); self.scene.render(p, QRectF(0, 0, rect.width(), rect.height()), rect); p.end()
            elif path.endswith('.dxf'): self._export_dxf(path)
        for i in hidden: i.show()
        QMessageBox.information(self, "完了", f"エクスポート完了:\n{path}")

    def _export_dxf(self, path):
        doc = ezdxf.new('R2010'); msp = doc.modelspace()
        for item in self.scene.items():
            is_preview = False
            try:
                if item == getattr(self.scene, 'preview_node', None) or item in getattr(self.scene, 'preview_items', []): is_preview = True
            except RuntimeError: pass
            if is_preview: continue
            
            if isinstance(item, NodeItem):
                x, y = item.scenePos().x(), -item.scenePos().y(); t = item.node_type
                if t == "process": coords = [(x-50, y+25), (x+50, y+25), (x+50, y-25), (x-50, y-25)]
                elif t == "decision": coords = [(x, y+35), (x+60, y), (x, y-35), (x-60, y)]
                elif t == "data": coords = [(x-35, y+25), (x+65, y+25), (x+35, y-25), (x-65, y-25)]
                elif t == "terminal": coords = [(x-35, y+25), (x+35, y+25), (x+50, y), (x+35, y-25), (x-35, y-25), (x-50, y)]
                else: coords = [(x-50, y+25), (x+50, y+25), (x+50, y-25), (x-50, y-25)]
                msp.add_lwpolyline(coords, close=True)
                if item.text_item.toPlainText():
                    ls = item.text_item.toPlainText().split('\n'); sy = y + (len(ls)-1)*7.5
                    for i, l in enumerate(ls): msp.add_text(l, dxfattribs={'height': 12}).set_placement((x, sy-i*15), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)
            elif isinstance(item, EdgeItem):
                pts = [item.source_node.scenePos()] + [wp.scenePos() for wp in item.waypoints] + [item.target_node.scenePos()]
                if len(pts) >= 2: pts[0] = clip_line_to_node(pts[0], pts[1], item.source_node); pts[-1] = clip_line_to_node(pts[-1], pts[-2], item.target_node)
                for i in range(len(pts)-1): msp.add_line((pts[i].x(), -pts[i].y()), (pts[i+1].x(), -pts[i+1].y()))
                if item.raw_text:
                    pos = item.get_auto_text_pos() + (item.text_item.manual_offset if item.text_item.manual_offset else QPointF(0,0))
                    mx, my = pos.x() + item.text_item.boundingRect().width()/2, -(pos.y() + item.text_item.boundingRect().height()/2)
                    ls = item.raw_text.split('\n'); sy = my + (len(ls)-1)*6
                    for i, l in enumerate(ls): msp.add_text(l, dxfattribs={'height': 10}).set_placement((mx, sy-i*12), align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER)
        doc.saveas(path)

    def copy_to_jwcad(self):
        td = ["JwcTemp", "hq", "lc7", "lt1"]
        for item in self.scene.items():
            is_preview = False
            try:
                if item == getattr(self.scene, 'preview_node', None) or item in getattr(self.scene, 'preview_items', []): is_preview = True
            except RuntimeError: pass
            if is_preview: continue
            
            if isinstance(item, NodeItem):
                x, y = item.scenePos().x(), -item.scenePos().y(); t = item.node_type
                def add_p(ps):
                    for i in range(len(ps)): td.append(f"{ps[i][0]} {ps[i][1]} {ps[(i+1)%len(ps)][0]} {ps[(i+1)%len(ps)][1]}")
                if t == "process": add_p([(x-50, y-25), (x+50, y-25), (x+50, y+25), (x-50, y+25)])
                elif t == "decision": add_p([(x, y-35), (x+60, y), (x, y+35), (x-60, y)])
                elif t == "data": add_p([(x-35, y-25), (x+65, y-25), (x+35, y+25), (x-65, y+25)])
                elif t == "terminal": add_p([(x-35, y-25), (x+35, y-25), (x+50, y), (x+35, y+25), (x-35, y+25), (x-50, y)])
                else: add_p([(x-50, y-25), (x+50, y-25), (x+50, y+25), (x-50, y+25)])
                if item.text_item.toPlainText():
                    ls = item.text_item.toPlainText().split('\n'); sy = y + (len(ls)-1)*7.5
                    for i, l in enumerate(ls): wc = sum(2 if unicodedata.east_asian_width(c) in 'FWA' else 1 for c in l); td.append(f'ch {x-wc*2.5} {sy-i*15-6.0} 10 0 "{l}')
            elif isinstance(item, EdgeItem):
                pts = [item.source_node.scenePos()] + [wp.scenePos() for wp in item.waypoints] + [item.target_node.scenePos()]
                if len(pts) >= 2: pts[0] = clip_line_to_node(pts[0], pts[1], item.source_node); pts[-1] = clip_line_to_node(pts[-1], pts[-2], item.target_node)
                for i in range(len(pts)-1): td.append(f"{pts[i].x()} {-pts[i].y()} {pts[i+1].x()} {-pts[i+1].y()}")
                if item.raw_text:
                    pos = item.get_auto_text_pos() + (item.text_item.manual_offset if item.text_item.manual_offset else QPointF(0,0))
                    mx, my = pos.x() + item.text_item.boundingRect().width()/2, -(pos.y() + item.text_item.boundingRect().height()/2)
                    ls = item.raw_text.split('\n'); sy = my + (len(ls)-1)*6
                    for i, l in enumerate(ls): wc = sum(2 if unicodedata.east_asian_width(c) in 'FWA' else 1 for c in l); td.append(f'ch {mx-wc*2.5} {sy-i*12-5.0} 10 0 "{l}')
        QApplication.clipboard().setText('\r\n'.join(td) + '\r\n')
        QMessageBox.information(self, "完了", "Jw_cad用のデータをクリップボードにコピーしました。\nJw_cadを開いて「編集」→「貼り付け (Ctrl+V)」を実行してください。")

    def open_print_dialog(self): 
        CustomPrintPreviewDialog(self, len(self.scene.selectedItems()) > 0).exec()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec())